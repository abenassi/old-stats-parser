#!C:\Python27
# -*- coding: utf-8 -*-
import sys
from openpyxl import load_workbook, Workbook
from utils import get_unicode, write_ws
import parsers
import stats_book_1


# INTERNAL CLASSES
class AbbyParser():
    """Parse rows of ABBY output and build records from them.

    AbbyParser object needs a list of parsers and a context to be built.
    Iterates through list of parsers looking for one that accepts the row, then
    parser modifies the context based on extracted information of row and
    AbbyParser build records from the context and return them."""

    def __init__(self, parsers, context, records_builder):
        self.context = context()
        self.parsers = parsers
        self.records_builder = records_builder

    def parse_row(self, row):
        """Main method. Parse a row modifying context an build records."""

        # iterate through parsers
        for parser_class in self.parsers:
            if parser_class(row).accepts():

                # parse row and modify context with results
                parser = parser_class(row, self.context)
                parser.parse()
                break

        # TODO: What happens if no parser accepts the row???

        # yield any new records that can be built, after row was parsed
        return self.records_builder(self.context).build_records()


# USER CLASSES
class AbbyFile():
    """Takes a workbook with a single sheet that is ABBY ocr output from an old
    stats book. Parse all rows of the worksheet building database records.

    Uses AbbyParser class to handle parsing wich needs parsers, context and
    record builder to do it."""

    def __init__(self, wb, parsers, context, record_builder, output_fields):
        self.wb = wb
        self.parsers = parsers
        self.context = context
        self.record_builder = record_builder
        self.output_fields = output_fields

    # PUBLIC
    def get_records(self):
        """Read all rows of workbook and parse them looking to build database
        records from them. Yield records as soon as they are built."""

        # load active sheet of wb
        ws = self.wb.get_active_sheet()

        # create AbbyParser instance
        ap = AbbyParser(self.parsers, self.context, self.record_builder)

        # iterate through all abby_file wb rows
        for row in ws.iter_rows():
            cells_values = []

            # iterate through all cells in row
            for cell in row:
                # add cell value
                cells_values.append(get_unicode(cell.internal_value))

            # checks if list of cell values not empty
            if not self._empty(cells_values):
                # remove any empty cell that might be at the end of row
                cells_values = self._remove_lasts_none(cells_values)

                # parse all posible records from row (list of cell values)
                record_lines = ap.parse_row(cells_values)

                # yields any record built from row
                for record in record_lines:
                    yield record

    # PRIVATE
    def _empty(self, values_list):
        """True if values_list is empty."""
        RV = True

        for value in values_list:
            if value is not None:
                RV = False

        return RV

    def _remove_lasts_none(self, values_list):
        """Delete None values at the end of a list."""
        RV = values_list

        while RV[-1] is None:
            del RV[-1]

        return RV


# DATA
ABBY_FILE_NAME = "abby_file.xlsx"
ABBY_PARSED_FILE_NAME = "abby_parsed.xlsx"

PARSERS = [parsers.IgnoreRow, parsers.NoneImportParser, parsers.Head1Parser,
           parsers.Head1IniPart, parsers.AgValuesParser, parsers.Head1FinalPart,
           parsers.TblRowParser, parsers.Head2Parser, parsers.TitleParser,
           parsers.Subt1Parser, parsers.Subt2Parser]

FIELDS = ["id_title",
          "desc_title",
          "id_subt1",
          "desc_subt1",
          "id_subt2",
          "desc_subt2",
          "id_product",
          "tariff_number",
          "desc_product",
          "product_units",
          "id_country",
          "desc_country",
          "year",
          "quantity",
          "value"]



def scrape_abby_file(wb_abby_name=None, wb_abby_parsed_name=None):
    """Takes an abby output excel file and returns a database formatted excel
    file with records built from it."""

    # if not wb names passed, defaults name are used
    wb_abby_name = wb_abby_name or ABBY_FILE_NAME
    wb_abby_parsed_name = wb_abby_parsed_name or ABBY_PARSED_FILE_NAME

    # loads abby file
    wb_abby = load_workbook(filename=wb_abby_name, use_iterators=True)
    abby_file = AbbyFile(wb_abby, PARSERS, stats_book_1.Context,
                         stats_book_1.RecordsBuilder, FIELDS)

    # creates new excel sheet to store new records
    wb_parsed = Workbook(optimized_write=True)
    ws_parsed = wb_parsed.create_sheet()

    # write field names
    ws_parsed.append(FIELDS)

    # write every record parsed in the database formatted excel sheet
    for record in abby_file.get_records():
        write_ws(ws_parsed, record, FIELDS)

    # save database formatted excel with parsed records
    wb_parsed.save(wb_abby_parsed_name)


# executes main routine
if __name__ == '__main__':

    # if parameters are passed, use them
    input_file = None
    output_file = None

    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]

    # main call
    scrape_abby_file(input_file, output_file)





