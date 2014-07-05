#!C:\Python27
# -*- coding: utf-8 -*-
from utils import get_unicode, find_nth, convert_to_float
from openpyxl import load_workbook, Workbook
import re

# DATA
ABBY_FILE_NAME = "abby_file.xlsx"


# ROW PARSERS
class BaseParser():

    def __init__(self, row=None, context=None):
        self.context = context
        self.row = row
        self.errors = [None]

    # @classmethod
    def accepts(self, row):
        """metodo que acepta una row en base a 3 condiciones.
        Si se proporciona el parametro para la condicion, esta
        se evalua. Si no, se asume como verdadera"""

        # carga los parametros de las condiciones en los objetos derivados
        self.load_conditions()

        # substring contained condition
        if self.id_string:
            contains_cond = self.id_string in row[0]
        else:
            contains_cond = True

        # length of row condition
        if self.row_length:
            len_cond = len(row) == self.row_length
        else:
            len_cond = True

        # pattern matching condition
        if self.row_pattern:
            pattern_cond = self._re_match(self.row_pattern, row[0])
        else:
            pattern_cond = True

        return contains_cond and len_cond and pattern_cond

    def _re_match(self, pattern, string):

        RV = False

        matchObj = re.match(pattern, string, re.U)

        if matchObj:
            RV = True

        return RV


class TitleParser(BaseParser):

    def load_conditions(self):
        self.id_string = u"T\xcdTULO"
        self.row_length = 1
        self.row_pattern = None

    def parse(self):
        self.context.nroTitulo = self._get_nroTitulo()
        self.context.descTitulo = self._get_descTitulo()

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "title"

    def _get_nroTitulo(self):

        i = self.row[0].find(".")

        RV = self.row[0][:i].replace(u'T\xcdTULO', "").strip()

        return RV

    def _get_descTitulo(self):

        i = self.row[0].find(".")
        pattern = "[A-Z][A-Z\s]{1,}"

        RV = re.search(pattern, self.row[0][i + 1:], re.U).group().strip()

        return RV


class Subt1Parser(BaseParser):

    def load_conditions(self):
        self.id_string = None
        self.row_length = 1
        self.row_pattern = "^[a-z][)][A-Z\s]{1,}"

    def parse(self):
        self.context.letraSubTitulo1 = self._get_letraSubTitulo1()
        self.context.descSubTitulo1 = self._get_descSubTitulo1()

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "subt1"

    def _get_letraSubTitulo1(self):

        i = self.row[0].find(")")

        RV = self.row[0][:i].strip()

        return RV

    def _get_descSubTitulo1(self):

        i = self.row[0].find(")")
        pattern = "[A-Z][A-Z\s]{1,}"

        RV = re.search(pattern, self.row[0][i + 1:], re.U).group().strip()

        return RV


class Subt2Parser(BaseParser):

    def load_conditions(self):
        self.id_string = None
        self.row_length = 1
        self.row_pattern = "^[0-9][\.]"

    def parse(self):
        self.context.nroSubTitulo2 = self._get_nroSubTitulo2()
        self.context.descSubTitulo2 = self._get_descSubTitulo2()

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "subt2"

    def _get_nroSubTitulo2(self):

        # print row
        i = self.row[0].strip().find(".")

        RV = self.row[0].strip()[:i]
        # print RV
        return RV

    def _get_descSubTitulo2(self):

        i = self.row[0].find(".")
        pattern = "[A-Z].{1,}"

        RV = re.search(pattern, self.row[0][i + 1:], re.U).group().strip()

        return RV


class AgValuesParser(BaseParser):

    def load_conditions(self):
        self.id_string = u"Valor total:"
        self.row_length = 1
        self.row_pattern = None

    def parse(self):

        # si la ultima row fue un titulo, pone en None los subtitulos
        if self.context.typeRow == "title":

            # Primer Subtitulo
            self.context.letraSubTitulo1 = 0
            self.context.descSubTitulo1 = "Todos"

            # Segundo Subtitulo
            self.context.nroSubTitulo2 = 0
            self.context.descSubTitulo2 = "Todos"

        # si la ultima row fue un titulo, pone en None el subt2
        if self.context.typeRow == "subt1":

            # Segundo Subtitulo
            self.context.nroSubTitulo2 = 0
            self.context.descSubTitulo2 = "Todos"

        # Tabla
        self.context.nroProducto = 0
        self.context.nroTarifa = u"NA"
        self.context.descProducto = u"Todos"
        self.context.unidadProducto = u"NA"

        # datos dentro de la Tabla que no se tienen
        self.context.idPais = 0
        self.context.descPais = "Todos"
        self.context.cantidad = [None, None]

        # datos de la tabla que se tienen
        self.context.year = self._get_year()
        self.context.value = self._get_value()

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "agValues"


    def _get_year(self):
        RV = []

        # agrega el primer año
        i = self.row[0].find(":")
        j = self.row[0].find("m$n")

        RV.append(int(self.row[0][i + 1:j].strip()))

        # agrega el segundo año
        i = self.row[0].find(";")
        j = self.row[0].rfind("m$n")

        RV.append(int(self.row[0][i + 1:j].strip()))

        return RV

    def _get_value(self):
        RV = []

        # agrega el valor del primer año
        i = self.row[0].find("m$n")
        j = self.row[0].find(";")

        strValue = self.row[0][i + 4:j]
        floatValue = convert_to_float(strValue)

        RV.append(floatValue)

        # agrega el valor del segundo año
        i = self.row[0].rfind("m$n")
        j = self.row[0].find(")")

        strValue = self.row[0][i + 4:j]
        floatValue = convert_to_float(strValue)

        RV.append(floatValue)

        return RV


class Head1Parser(BaseParser):

    def load_conditions(self):
        self.id_string = "Tarifa"
        self.row_length = 1
        self.row_pattern = "^[0-9].{1,}Tarifa.{1,}:"

    def parse(self):

        # extrae los datos que devuelve el parser
        self.context.nroProducto = self._get_nroProducto()
        self.context.nroTarifa = self._get_nroTarifa()
        self.context.descProducto = self._get_descProducto()
        self.context.unidadProducto = self._get_unidadProducto()

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "tblHead1"

    def _get_nroProducto(self):

        indexList = [self.row[0].strip().find("."),
                     self.row[0].strip().find("-"),
                     self.row[0].strip().find("(")]

        # remueve el -1, si existe
        try:
            indexList.remove(-1)
        except:
            pass

        # calcula el indice minimo
        minIndex = min(indexList)

        # si es mayor que uno, lo utiliza
        if minIndex > 1:
            i = minIndex
        # si no, utiliza 1 como minimo absoluto
        else:
            i = 1

        RV = self.row[0].strip()[:i]

        return RV

    def _get_nroTarifa(self):

        if u"varios y no tarifados" in self.row[0]:
            RV = "varios y no tarifados"

        else:

            iStart = self.row[0].find(u"Tarifa") + 6
            # iEnd = find_nth(self.row[0], ".", 2) - 1
            iEnd = self.row[0].find(u")")

            RV = self.row[0][iStart:iEnd].strip()

        return RV

    def _get_descProducto(self):

        ### chequea que haya un "primer punto" antes de los parentesis
        # busca el primer punto que encuentra
        iDot = self.row[0].find(".")

        # busca los indices de los parentesis
        iParenthesis = [self.row[0].find("("), self.row[0].find(")")]

        # remueve el -1, si existe
        try:
            iParenthesis.remove(-1)
        except:
            pass

        # calcula el indice minimo
        minIndex = min(iParenthesis)

        # si el indice del primer punto es menor, esta antes de los parentesis
        if iDot < minIndex:
            iStart = find_nth(self.row[0], ".", 2) - 1

        # si no, hay que tomar entonces el primer punto (ya que falta el otro)
        else:
            iStart = find_nth(self.row[0], ".", 1) - 1

        # busca el indice final de la descripcion a partir de la ultima coma
        iEnd = self.row[0].rfind(",")

        # regex pattern
        pattern = "[A-Z].{1,}"

        # intenta matchear el patron, sino devuelve error
        try:
            RV = re.search(pattern, self.row[0][iStart:iEnd], re.U).group().strip()
        except:
            RV = "Parsing error"

        return RV

    def _get_unidadProducto(self):

        iStart = self.row[0].rfind(",") + 1
        iEnd = self.row[0].rfind(":")

        RV = self.row[0][iStart:iEnd].strip()
        RV = self._homogeneizar_unidadProducto(RV)

        return RV

    def _homogeneizar_unidadProducto(self, unidadProducto):

        RV = unidadProducto

        if get_unicode(unidadProducto.strip()) == u"Kg.":
            RV = u"kilogramos"

        return RV


class Head2Parser(BaseParser):

    def load_conditions(self):
        self.id_string = None
        self.row_length = 1
        self.row_pattern = "^[0-9].{1,}:"


class TblRowParser(BaseParser):

    def load_conditions(self):
        self.id_string = None
        self.row_length = 5
        self.row_pattern = None

    def parse(self):

        # datos dentro de la Tabla
        self.context.idPais = None
        self.context.descPais = self._get_descPais()
        self.context.year = self._get_year()
        self.context.cantidad = self._get_cantidad()
        self.context.value = self._get_value()

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "tblRow"

    def _get_descPais(self):

        if not self.row[0]:
            RV = "Missing error"

        elif u"Total" in self.row[0] or u"total" in self.row[0]:
            RV = u"Todos"

        else:
            RV = self.row[0].replace(".", "").strip()

        return RV

    def _get_year(self):

        RV = [u"1945", u"1946"]

        return RV

    def _get_cantidad(self):

        cant_1945 = None
        cant_1946 = None

        try:
            cant_1945 = convert_to_float(self.row[1])
        except:
            pass

        try:
            cant_1946 = convert_to_float(self.row[2])
        except:
            pass

        RV = [cant_1945, cant_1946]

        return RV

    def _get_value(self):

        value_1945 = None
        value_1946 = None

        try:
            value_1945 = convert_to_float(self.row[3])
        except:
            pass

        try:
            value_1946 = convert_to_float(self.row[4])
        except:
            pass

        RV = [value_1945, value_1946]

        return RV


class IgnoreRow(BaseParser):

    # @classmethod
    def accepts(self, row):
        """pisa el metodo base, chequea condiciones especiales
        para las rows que deben ser ignoradas"""

        # si la row es None en su primer celda, se ignora
        if (not row) or (not row[0]):
            return True

        # chequea si la row es una continuacion de tabla partida
        continuation_row = u"(Conclusi\xf3n)" in row[0]

        # chequea si la row esta vacia
        empty_row = u"" == row[0]

        return continuation_row or empty_row

    def parse(self):

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "ignore"


class NoneImportParser(BaseParser):

    def load_conditions(self):
        self.id_string = u"Sin importaci\xf3n"
        self.row_length = 1
        self.row_pattern = None

    def parse(self):

        # datos dentro de la Tabla
        self.context.idPais = 0
        self.context.descPais = u"Todos"
        self.context.year = [1945, 1946]
        self.context.cantidad = [u"NA", u"NA"]
        self.context.value = [u"NA", u"NA"]

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "noneImport"

class Head1IniPart(BaseParser):

    def load_conditions(self):
        self.id_string = None
        self.row_length = 1
        self.row_pattern = "^[0-9].{1,}Tarifa.{1,}"

    def parse(self):
        """toma la parte inicial de un head1 y la guarda hasta que
        aparezca la parte final"""

        self.context.lastRow = self.row[0]

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "tblHead1IniPart"


class Head1FinalPart(BaseParser):

    def load_conditions(self):
        self.id_string = None
        self.row_length = 1
        self.row_pattern = ".{1,}:"

    def parse(self):
        """toma la parte final de un head1 y la concatena con la parte
        inicial previamente guardada, luego llama al metodo de head1"""

        # forma la row completa de tipo Head1
        tblHead1Row = [get_unicode(self.context.lastRow +
                       u" " + self.row[0])]
        tblHead1Row = [tblHead1Row[0].strip()]

        # chequea que sea aceptada, y usa el parser de Head1
        if Head1Parser().accepts(tblHead1Row):
            Head1Parser(tblHead1Row, self.context).parse()
        else:
            print tblHead1Row
            print "Ocurrio un error con un Head1Parser partido!"

        # declara cual fue el ultimo row type procesado
        self.context.typeRow = "tblHead1FinalPart"

# INTERNAL CLASSES
class Anuario1Context():

    def __init__(self):

       # Titulo
        self.nroTitulo = None
        self.descTitulo = None

        # Primer Subtitulo
        self.letraSubTitulo1 = None
        self.descSubTitulo1 = None

        # Segundo Subtitulo
        self.nroSubTitulo2 = None
        self.descSubTitulo2 = None

        # Tabla
        self.nroProducto = None
        self.nroTarifa = None
        self.descProducto = None
        self.unidadProducto = None

        # Datos dentro de la Tabla
        self.idPais = None
        self.descPais = None
        self.year = []
        self.cantidad = []
        self.value = []

        # cursor
        self.typeRow = ""
        self.typeLastRow = ""
        self.lastRow = ""


class AbbyParser():

    def __init__(self, parsers, context):

        # contexto y parsers
        self.context = context()
        self.parsers = parsers

        # resultados
        self.errors = []

    def parse_row(self, row):

        # recorre los parsers a ver si alguno acepta la row
        for parserClass in self.parsers:
            if parserClass().accepts(row):
                # print str(parserClass), row
                # parsea la row y modifica el contexto con el resultado
                parser = parserClass(row, self.context)
                parser.parse()
                break

        # FALTA IMPLEMENTAR CUANDO NO SE RECONOCE EL TIPO!!!
        # devuelve los registros que se puedan construir
        return self._build_records()

    def get_errors(self):
        """devuelve la lista de errors y la deja vacia"""

        RV = list(self.errors)
        self.errors = []

        return RV

    # METODOS PRIVADOS
    def _build_records(self):

        newRecords = []

        if self.context.typeRow == "agValues" or \
                self.context.typeRow == "tblRow":

            i = 0
            for value in self.context.value:

                newRecord = dict()

                # Titulo
                newRecord["nroTitulo"] = self.context.nroTitulo
                newRecord["descTitulo"] = self.context.descTitulo

                # Primer Subtitulo
                newRecord["letraSubTitulo1"] = self.context.letraSubTitulo1
                newRecord["descSubTitulo1"] = self.context.descSubTitulo1

                # Segundo Subtitulo
                newRecord["nroSubTitulo2"] = self.context.nroSubTitulo2
                newRecord["descSubTitulo2"] = self.context.descSubTitulo2

                # Tabla
                newRecord["nroProducto"] = self.context.nroProducto
                newRecord["nroTarifa"] = self.context.nroTarifa
                newRecord["descProducto"] = self.context.descProducto
                newRecord["unidadProducto"] = self.context.unidadProducto

                # datos dentro de la Tabla
                newRecord["idPais"] = self.context.idPais
                newRecord["descPais"] = self.context.descPais
                newRecord["year"] = self.context.year[i]
                newRecord["cantidad"] = self.context.cantidad[i]
                newRecord["value"] = value

                # agrega el nuevo record
                newRecords.append(newRecord)
                i += 1

        # actualiza el ultimo tipo de row tratado
        self.context.typeRow = str(self.context.typeRow)

        return newRecords


# USER CLASSES
class AbbyFile():

    # DATA
    PARSERS = [IgnoreRow, NoneImportParser, Head1Parser,
               Head1IniPart,
               AgValuesParser, Head1FinalPart,
               TblRowParser, Head2Parser,
               TitleParser, Subt1Parser, Subt2Parser]

    CONTEXTS = [Anuario1Context]

    FIELDS = ["nroTitulo",
              "descTitulo",
              "letraSubTitulo1",
              "descSubTitulo1",
              "nroSubTitulo2",
              "descSubTitulo2",
              "nroProducto",
              "nroTarifa",
              "descProducto",
              "unidadProducto",
              "idPais",
              "descPais",
              "year",
              "cantidad",
              "value"]

    def __init__(self, wb):
        self.wb = wb
        self.errors = []

    def get_records(self):

        # toma la hoja activa
        ws = self.wb.get_active_sheet()

        # crea un abby parser
        ap = AbbyParser(self.PARSERS, self.CONTEXTS[0])

        # lee optimizadamente todo el archivo
        for row in ws.iter_rows():

            # crea lista vacia que tendra los values de la fila leida
            values_list = []

            # recorre las celdas de la fila leida
            for cell in row:

                # agrega value de la celda a la lista de la row
                values_list.append(get_unicode(cell.internal_value))

            # si la lista no esta vacia, la parsea
            if not self._empty(values_list):

                # elimina todas las celdas vacias del final de la lista
                values_list = self._remove_lasts_none(values_list)
                # print values_list
                # extrae los records que se pueda a partir de la row pasada
                record_lines = ap.parse_row(values_list)
                # print record_lines
                # devuelve de a uno cada record nuevo conseguido a medida
                # que se generan nuevos records con las rows leidas del excel
                for record in record_lines:
                    yield record

    # PRIVATE
    def _empty(self, valuesList):
        """indica si una lista esta vacia"""

        RV = True

        for value in valuesList:
            if value != None:
                RV = False

        return RV

    def _remove_lasts_none(self, valuesList):
        """elimina todos los None values del final de una lista"""

        RV = valuesList

        while RV[-1] is None:
            del RV[-1]

        return RV


# USER METHODS
def write_ws(ws, record, fields):
    """recibe una hoja de excel vacia o que ya tiene algunos
    registros y debe agregar un nuevo registro"""

    new_row = []
    # print record
    # agrega en orden los valores segun el field al que pertenecen
    for field in fields:
        new_row.append(record[field])

    # agrega la nueva row al excel
    ws.append(new_row)


def scrape_abby1File(wb_abby_name=None):
    """toma un excel con tablas del abby y devuelve un excel con una
    tabla en formato de base de datos"""

    # si no se pasa un nombre de archivo, se toma del modulo
    wb_abby_name = wb_abby_name or ABBY_FILE_NAME

    # carga el archivo
    wb_abby = load_workbook(filename=wb_abby_name, use_iterators=True)

    # creo un objeto abby
    abby_file = AbbyFile(wb_abby)

    # creo una hoja de excel para ir guardando el output
    wb_parsed = Workbook(optimized_write=True)
    ws_parsed = wb_parsed.create_sheet()

    # copio los encabezados
    ws_parsed.append(abby_file.FIELDS)

    # reescribe cada record en un nuevo excel con formato de base de datos
    for record in abby_file.get_records():
        write_ws(ws_parsed, record, abby_file.FIELDS)

    # guarda el excel terminado
    wb_parsed.save("ABBY parsed refactored.xlsx")

    return abby_file


"""
    1. Implementar lo que pasa cuando no se reconoce el tipo
    2. Implementar la rutina que va agregando records a la hoja
    de excel
    3. Implementar algun tipo de registro y reporte de errores
    4. Implementar alguna logica "anti-abby-errors" para los
    valores numericos en donde se cuela alguna letra
    5. Implementar homogeneizacion de nombres de paises
"""






